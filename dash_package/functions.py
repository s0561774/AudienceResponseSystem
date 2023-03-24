from pandas import DataFrame
from openpyxl import load_workbook
import uuid
from datetime import date
import numpy as np
from dash_package.Pool import Pool


def saveQuestion(questions):
    print(questions)
    num = uuid.uuid4().__str__()
    li = [date.today().isoformat(),num]
    ans=[num]
    str=''
    correct_response=''
    pool = Pool(num)
    wb = load_workbook('app_questions.xlsx')
    ws = wb.worksheets[1]
    ws.append(ans)
    for x in range(2, len(questions)):
        myArray= questions[x][1][0].split('|')
        str +=myArray[0] + "|"
        if(len(myArray)>1):
            correct_response +=myArray[0] + "|"
        pool.questions.append(myArray[0])
        print(myArray)
    for q in questions[:2]:
        li.append(q[1][0])
    li.append(str)
    li.append(0)
    li.append(correct_response)
    ws = wb.worksheets[0]
    ws.append(li)
    wb.save('app_questions.xlsx')
    return pool


def getDatas():
    res = []
    wb = load_workbook('app_questions.xlsx')
    ws = wb.worksheets[0]
    max_col = ws.max_column
    max_row = ws.max_row
    for j in range(2, max_row+1):
        curr_res=[]
        for i in range(1, max_col + 1):
            curr_res.append(ws.cell(row = j, column = i).value)

        res.append(curr_res)

    results = []
    test_list = []
    for l in res:
        l = list(filter(lambda x: x is not None, l))
        test_list.append(l)

    results = list(filter(None, test_list))
    return results


def startPool(id_question):
    wb = load_workbook('app_questions.xlsx')
    ws = wb.worksheets[0]
    max_row = ws.max_row
    max_row +=1
    print(id_question)
    change = False
    for j in range(2, max_row):
        if ws.cell(row= j, column= 2).value == id_question:
            print('entrer')
            ws.cell(row = j, column = 6).value = 1
            change = True
    
    wb.save('app_questions.xlsx')
    return change
    
    
def stopPool(id_question):
    wb = load_workbook('app_questions.xlsx')
    ws = wb.worksheets[0]
    max_row = ws.max_row
    max_row +=1
    change = False
    for j in range(2, max_row):
        if ws.cell(row= j, column= 2).value == id_question:
            ws.cell(row = j, column = 6).value= 0 
            change = True
    
    wb.save('app_questions.xlsx')
    return change

def findQuestion(question_id):
    wb = load_workbook('app_questions.xlsx')
    ws = wb.worksheets[0]
    max_col = ws.max_column
    max_row = ws.max_row
    max_row +=1
    frage = []
    found = False
    for j in range(2, max_row):
        if ws.cell(row= j, column= 2).value == question_id:
            found = True
            for i in range(2, max_col + 1):
                frage.append(ws.cell(row = j, column = i).value)

        if  found == True:
            break        
    return frage

def optionselected(question_id,option):
    wb = load_workbook('app_questions.xlsx')
    ws_1 = wb.worksheets[1]
    ws_0 = wb.worksheets[0]
    max_row_0 = ws_0.max_row
    max_row_1 = ws_1.max_row
    change = False
    isStarted= False
    max_row_0 +=1
    max_row_1 +=1
    for j in range(2, max_row_0):
        if ws_0.cell(row= j, column= 2).value == question_id:
            if ws_0.cell(row = j, column = 6).value == 1:
                isStarted = True
                break
    if isStarted == True:
        for j in range(2, max_row_1):
            if ws_1.cell(row= j, column= 1).value == question_id:
                ws_1.cell(row = j, column = option).value= int(ws_1.cell(row = j, column = option).value or 0) +1 
                change = True
    
    wb.save('app_questions.xlsx')
    return change


def getResponse():
    res = []
    wb = load_workbook('app_questions.xlsx')
    ws = wb.worksheets[1]
    max_col = ws.max_column
    max_row = ws.max_row
    
    for j in range(2, max_row+1):
        curr_res=[]
        for i in range(1, max_col + 1):
            curr_res.append(ws.cell(row = j, column = i).value)
        
        res.append(curr_res)
    results = []
    test_list = []
    for l in res:
        l = list(filter(lambda x: x is not None, l))
        if len(l)>1:
            test_list.append(l)

    results = list(filter(None, test_list))
    return results

def get_question_for_chart() -> list:
    data_response = getResponse()
    data_questions =  getDatas()
    data_for_chart= list()
    for q in data_questions:
        for r in data_response:
            if q[1] == r[0]:
                data_for_chart.append(q)

    return data_for_chart


def get_data_for_chart(id_question)->list:
    dt_ch = get_question_for_chart()
    dt_r = getResponse()
    results =list()
    for res in dt_ch:
        if res[1] == id_question:
            options = list(filter(lambda x: x != '', res[4].split('|') ))
            results.append(options)
            results.append(list(filter(lambda x: x != '', res[6].split('|') )))
            break
    for  val in dt_r:
        if val[0] == id_question:
            data =val[1:]
            print(data)
            results.append(data)
            break
    
    while len(results[0]) > len(results[2]):
        results[2].append(0)

    return results


def remove(question_id):
    wb = load_workbook('app_questions.xlsx')
    ws_0 = wb.worksheets[0]
    ws_1 = wb.worksheets[1]
    max_row_0 = ws_0.max_row
    max_row_1 = ws_1.max_row
    max_row_0 +=1
    max_row_1 +=1
    found = False
    for j in range(2, max_row_0):
        if ws_0.cell(row= j, column= 2).value == question_id:
            ws_0.delete_rows(j,1)
            found = True
            break

    for j in range(2, max_row_1):
            if ws_1.cell(row= j, column= 1).value == question_id:
                ws_1.delete_rows(j,1)
                break
    wb.save('app_questions.xlsx')
    print(found)
    return found
  